#! python
# updateProduce.py - corrige os preços em uma planilha de venda de produtos.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# os tipos de produtos e seus preços atualizdos

PRICE_UPDATES = {'Garlic': 5.09,
                 'Celery': 3.19,
                 'Lemon': 1.07}

# Percorre as linhas em um loop e atualiza os preços

for rowNum in range(2, sheet.max_row):  # pula a primeira linha
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
           sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('produceSalesnovo.xlsx')
